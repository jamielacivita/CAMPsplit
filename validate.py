#####################################################################################################################################################################################################
    #CAMP Input Validator
    #Input : .xlsx file modified for anticipated input into the CAMP system.
    #Output: Text indicating if data format of input file is correct as required for CAMP Data Migration Process
#####################################################################################################################################################################################################


from openpyxl import Workbook, load_workbook
import datetime
import sys
import argparse

def checkColumn(ws, checkColumn, expectedValue):
    """Given a column number and an expected value return true/false indicating if that value is present at column"""
    returnValue = False
    if ws.cell(row=1, column=checkColumn).value == expectedValue:
        returnValue = True
    return returnValue

def verifyValue(ws, column, value, verbose):
    cellHasValidHeading = True
    if checkColumn(ws, column,value):
        if (verbose) : print("Column " + numberToLetter(column) + " is correct.")
        pass
    else:
        cellHasValidHeading = False
        #print("Column " + numberToLetter(column) + " is not correct!  Expected " + value)
        print("Checked {:50s} {:30s} expected '{}'".format(numberToLetter(column), "\033[91m...FAILED\033[0m",value))
    return cellHasValidHeading

def numberToLetter(colNumber):
    """Given a column number return the excel column heading"""
    ascii = colNumber + 64
    return chr(ascii)

def validateColEntries(ws, colNumber, rowNumber, validEntriesLst):
    """Verify for a given col, row the entry is in the validEntries list"""
    returnValue = False
    cellValue = ws.cell(row=rowNumber, column=colNumber).value
    if cellValue in validEntriesLst:
        returnValue = True
    return returnValue

def checkColumnData(ws, max_rows, colName, colNumber):
    """Given a column name and a column number performs checking for valid entries on that column"""
    #print("Checking " + str(colName))
    allValidData = True
    for n in range(2,max_rows+1):
        if (validateColEntries(ws, colNumber,n,valColEntries[colNumber])):
            pass
        else:
            allValidData = False
            #print("In " + str(colName) +" Row " + str(n) + " is bad \t\t\t")
            print("Checked {:50s} {:30s}".format(colName, "\033[91m...FAILED\033[0m"))
    if (allValidData==True):
        print("Checked {:50s} {:30s}".format(colName, "...passed"))


def setFilename(fn):
    filename = fn


def verifyHeadings(ws, verbose):
    print("Validating column header row...")
    headingsAreValid = True  #presumed all headings are correct. 
    #Verify Each column headings is correct
    for n in range(1,26):
        if(verifyValue(ws, n,colHeadings[n], verbose)==False): 
            allColumnHeadingsValid = False
    if (headingsAreValid):
        print("Column Header Validation Passed.\n")
    else:
        print("Column Header Validation Failed.\n")
    return headingsAreValid


def allColumnHeadingsGood(ws,verbose):
    allColumnHeadingsValid = True  #presumed all headings are correct. 
    for n in range(1,26):
        returnValue = verifyValue(ws, n,colHeadings[n], verbose)
        if (returnValue == False):
            allColumnHeadingsValid = False
    return allColumnHeadingsValid


def verifyColumnData(ws, verbose):
    #Iterate over the list of column numbers to check and pass in the name of the colum (dictionary lookup) and the column number.
    for c in colNumToCheck:
        checkColumnData(ws, ws.max_row, colHeadings[c], c)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("--verbose", action='store_true')
    args = parser.parse_args()

    #load workbook
    print("\nValidating: " + args.filename)
    wb = load_workbook(args.filename)

    #activate sheet
    ws = wb.active
    max_rows = ws.max_row

    verifyHeadings(ws, args.verbose)
    verifyColumnData(ws, args.verbose)


#ColHeadings is a dictionary that stores key value pair of column number and heading title.
#Example: Column 1 is titled "Entity Name"
colHeadings = {}
colHeadings[1] = "Entity Name"
colHeadings[2] = "Entity Unique ID"
colHeadings[3] = "Legacy 432 Entity ID"
colHeadings[4] = "External Entity ID"
colHeadings[5] = "Alias"
colHeadings[6] = "Sourcing Company"
colHeadings[7] = "Entity Country"
colHeadings[8] = "Entity Risk Rating"
colHeadings[9] = "Access Risk Rating"
colHeadings[10] = "Competitor"
colHeadings[11] = "Subject to Export Compliance Laws"
colHeadings[12] = "Contractual or Local Law Restrictions"
colHeadings[13] = "High Risk Country"
colHeadings[14] = "Date of Last Review"
colHeadings[15] = "Entity Info Type: IP Access"
colHeadings[16] = "Entity Info Type:  SPD Access"
colHeadings[17] = "Entity Info Type:  TD Access"
colHeadings[18] = "Data Info Classification"
colHeadings[19] = "Access without a Chevron ID:"
colHeadings[20] = "Access without a Chevron ID:  Additional Guidance"
colHeadings[21] = "Access with a Chevron ID:"
colHeadings[22] = "Access with a Chevron ID:  Additonal Guidance"
colHeadings[23] = "Email Access:"
colHeadings[24] = "Shared Drive:"
colHeadings[25] = "Intranet:"


#valColEntries is a dictionary that holds a list of valid entries for a (key) colum number.
#Example : Col 8 can hold either "High Risk" or "Low Risk" as valid entries.
valColEntries = {}
valColEntries[8] = ["High Risk", "Low Risk"]
valColEntries[9] = ["High Risk Access", "Low Risk Access"]
valColEntries[10] = ["Yes", "No"]
valColEntries[11] = ["Yes", "No"]
valColEntries[12] = ["Yes", "No"]
valColEntries[13] = ["Yes", "No"]
valColEntries[15] = ["Yes", "No"]
valColEntries[16] = ["Yes", "No"]
valColEntries[17] = ["Yes", "No"]
valColEntries[18] = ["None", "Classified", "Confidential-Restricted Access", "Company Confidential"]
valColEntries[19] = ["Yes", "No"]
valColEntries[20] = ["Yes", "No"]
valColEntries[21] = ["Yes", "No"]
valColEntries[22] = ["Yes", "No"]
valColEntries[23] = ["Yes", "No"]
valColEntries[24] = ["Yes", "No"]
valColEntries[25] = ["None", "Basic", "Full"]

#colNumToCheck is a list of column numbers against which the checkColumnData method should be applied. 
colNumToCheck = [8,9,10,11,12,13,15,18,19,20,21,22,23,24,25]


log = True

if __name__ == "__main__":
    main()

