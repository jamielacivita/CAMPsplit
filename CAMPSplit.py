#####
#Input : CAMP Data File
#Output : That same file split into chunks
#####

import argparse
from openpyxl import Workbook, load_workbook
import validate

def getLineSplits(max_row, chunkSize):
    """Return a list of tuples showing where the lines should break."""
    lineSplits = []

    start = 2 #presume header row
    while(start <= max_row):
        end = start + chunkSize-1
        if (end >= max_row):
            lineSplits.append((start,max_row))
            break
        lineSplits.append((start,end))
        start = end + 1

    return lineSplits

def getFileLetter(fileNumber):
    return chr(64+fileNumber)

def createOutputFile(inputSheet, fileNumber,startRow,endRow,max_column,max_row,outputFileName):
    """Creates an output file with a heading row and rows from startRow to endRow from inputFile"""
    #create output sheet
    outputBook = Workbook()
    outputSheet = outputBook.active
    #outputSheet['A1'] = "JWTO"
    #copy in header
    for col in range(1,max_column+1):
        #read in data from input sheet
        value = inputSheet.cell(row=1,column=col).value
        #print("Read Value: " + value)
        outputSheet.cell(row=1,column=col).value = value
    #copy in rows
    outRow = 2
    for r in range(startRow,endRow+1):
        for c in range(1,max_column+1):
            #read in data from input sheet
            value = inputSheet.cell(row=r,column=c).value
            #output the value to the new sheet.
            outputSheet.cell(row=outRow,column=c).value = value
        outRow = outRow+1

    #save the output file
    outputBook.save(outputFileName+"_"+getFileLetter(fileNumber)+".xlsx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("inputFilename")
    parser.add_argument("outputFilename")
    parser.add_argument("chunkSize", type=int)
    args = parser.parse_args()

    showSplits = True
    log = True
    createFiles = True
    validateHeadings = True
    validateColumns = True

    #Read in input file
    inputBook = load_workbook(args.inputFilename)

    inputSheet = inputBook.active

    #Get max row count
    max_row = inputSheet.max_row

    #get max column count
    max_column = inputSheet.max_column

    if(validateHeadings):
        print("Verifying Column Headings")
        #print("Result of verifyHeadings: " + str(validate.verifyHeadings(inputSheet, max_row)))
        goodHeader = validate.verifyHeadings(inputSheet, max_row)
        if(not goodHeader):
            print("Header failed validation, files will not be created.")

    if(validateColumns):
        print("Verifying Column Data")
        #print("Results of verifyColumnData: " + str(validate.verifyColumnData(inputSheet, max_row)))
        goodData = validate.verifyColumnData(inputSheet, max_row)
        if(not goodData):
            print("Data failed validation, files will not be created.")

    if (createFiles and goodHeader and goodData):
        print("Creating output files")
        splits = getLineSplits(max_row,args.chunkSize)
        outputFileNumber = 1
        for c in splits:
            #print(str(c[0]) + " " + str(c[1]))
            createOutputFile(inputSheet, outputFileNumber,c[0],c[1],max_column,max_row,args.outputFilename)
            outputFileNumber = outputFileNumber + 1

main()
